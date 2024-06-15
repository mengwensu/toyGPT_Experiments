import os
import pickle
from contextlib import nullcontext
import torch
import tiktoken
from model import GPTConfig, GPT
import numpy as np
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics.pairwise import cosine_similarity
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import Workbook, load_workbook

import plotly.express as px
import plotly.io as pio
from scipy.signal import correlate




# -----------------------------------------------------------------------------
init_from = 'resume' # either 'resume' (from an out_dir) or a gpt2 variant (e.g. 'gpt2-xl')
out_dir = 'out' # ignored if init_from is not 'resume'
start = "\n" # or "" or etc. Can also specify a file, use as: "FILE:prompt.txt"
num_samples = 10 # number of samples to draw
max_new_tokens = 500 # number of tokens generated in each sample
temperature = 1.2 # 1.0 = no change, < 1.0 = less random, > 1.0 = more random, in predictions
top_k = 200 # retain only the top_k most likely tokens, clamp others to have 0 probability
seed = 1337
device = 'cuda' # examples: 'cpu', 'cuda', 'cuda:0', 'cuda:1', etc.
dtype = 'bfloat16' if torch.cuda.is_available() and torch.cuda.is_bf16_supported() else 'float16' # 'float32' or 'bfloat16' or 'float16'
compile = False # use PyTorch 2.0 to compile the model to be faster
exec(open('configurator.py').read()) # overrides from command line or config file
# -----------------------------------------------------------------------------

torch.manual_seed(seed)
torch.cuda.manual_seed(seed)
torch.backends.cuda.matmul.allow_tf32 = True # allow tf32 on matmul
torch.backends.cudnn.allow_tf32 = True # allow tf32 on cudnn
device_type = 'cuda' if 'cuda' in device else 'cpu' # for later use in torch.autocast
ptdtype = {'float32': torch.float32, 'bfloat16': torch.bfloat16, 'float16': torch.float16}[dtype]
ctx = nullcontext() if device_type == 'cpu' else torch.amp.autocast(device_type=device_type, dtype=ptdtype)

# model
if init_from == 'resume':
    # init from a model saved in a specific directory
    ckpt_path = os.path.join(out_dir, 'ckpt.pt')
    checkpoint = torch.load(ckpt_path, map_location=device)
    gptconf = GPTConfig(**checkpoint['model_args'])
    model = GPT(gptconf)
    state_dict = checkpoint['model']
    unwanted_prefix = '_orig_mod.'
    for k, v in list(state_dict.items()):
        if k.startswith(unwanted_prefix):
            state_dict[k[len(unwanted_prefix):]] = state_dict.pop(k)
    model.load_state_dict(state_dict)
elif init_from.startswith('gpt2'):
    # init from a given GPT-2 model
    model = GPT.from_pretrained(init_from, dict(dropout=0.0))

model.eval()
model.to(device)
if compile:
    model = torch.compile(model)  # requires PyTorch 2.0 (optional)

# look for the meta pickle in case it is available in the dataset folder
load_meta = False
if init_from == 'resume' and 'config' in checkpoint and 'dataset' in checkpoint['config']:  # older checkpoints might not have these...
    meta_path = os.path.join('data', checkpoint['config']['dataset'], 'meta.pkl')
    load_meta = os.path.exists(meta_path)
if load_meta:
    print(f"Loading meta from {meta_path}...")
    with open(meta_path, 'rb') as f:
        meta = pickle.load(f)
    # TODO want to make this more general to arbitrary encoder/decoder schemes
    stoi, itos = meta['stoi'], meta['itos']
    encode = lambda s: [stoi[c] for c in s]
    decode = lambda l: ''.join([itos[i] for i in l])
else:
    # ok let's assume gpt-2 encodings by default
    print("No meta.pkl found, assuming GPT-2 encodings...")
    enc = tiktoken.get_encoding("gpt2")
    encode = lambda s: enc.encode(s, allowed_special={""})
    decode = lambda l: enc.decode(l)

# encode the beginning of the prompt
# if start.startswith('FILE:'):
#     with open(start[5:], 'r', encoding='utf-8') as f:
#         start = f.read()
# start_ids = encode(start)
# # start_ids = encode("ABCD")
# x = (torch.tensor(start_ids, dtype=torch.long, device=device)[None, ...])
# print(x)



def calculate_angle(vector1, vector2):
    # Convert to numpy arrays if they are not already
    vector1 = np.array(vector1)
    vector2 = np.array(vector2)
    
    # Compute the dot product
    dot_product = np.dot(vector1, vector2)
    
    # Compute the magnitudes (norms) of each vector
    norm1 = np.linalg.norm(vector1)
    norm2 = np.linalg.norm(vector2)
    
    # Compute the cosine similarity
    cosine_similarity = dot_product / (norm1 * norm2)
    
    # Clip values to avoid numerical issues with arccos
    cosine_similarity = np.clip(cosine_similarity, -1.0, 1.0)
    
    # Compute the angle in radians and then convert to degrees
    angle_radians = np.arccos(cosine_similarity)
    angle_degrees = np.degrees(angle_radians)
    
    return angle_degrees

def calculate_angle_matrix(embeddings):
    num_vectors = len(embeddings)
    angle_matrix = np.zeros((num_vectors, num_vectors))
    
    for i in range(num_vectors):
        for j in range(i, num_vectors):  # Only compute upper triangular matrix to avoid redundancy
            angle = calculate_angle(embeddings[i], embeddings[j])
            angle_matrix[i, j] = angle
            angle_matrix[j, i] = angle  # Symmetric matrix
    return angle_matrix

def compute_cross_correlation(vec1, vec2):
    # Example function to compute cross-correlation
    return np.correlate(vec1, vec2, mode='full').max()

def normalize_cross_correlation(cross_corr, norm1, norm2):
    # Example function to normalize cross-correlation
    return cross_corr / (norm1 * norm2)

def cross_correlation_distance(embeddings):
    num_vectors = len(embeddings)
    distance_matrix = np.zeros((num_vectors, num_vectors))
    
    norms = [np.linalg.norm(vec) for vec in embeddings]
    
    for i in range(num_vectors):
        for j in range(i, num_vectors):
            cross_corr = compute_cross_correlation(embeddings[i], embeddings[j])
            normalized_corr = normalize_cross_correlation(cross_corr, norms[i], norms[j])
            # Transform similarity to distance (1 - similarity)
            distance = 1 - normalized_corr
            distance_matrix[i, j] = distance
            distance_matrix[j, i] = distance
    
    # print("dm_shape: ", distance_matrix.shape)
    # Extract the upper triangular part of the distance matrix excluding the diagonal
    upper_triangular_indices = np.triu_indices(num_vectors, k=1)
    # print("indices:", upper_triangular_indices)
    upper_triangular_values = distance_matrix[upper_triangular_indices]
    # print("triangle:", upper_triangular_values)
    x_corr = np.mean(upper_triangular_values)
    
    return x_corr

def write_matrix_to_excel(angle_matrix, filename):
    df = pd.DataFrame(angle_matrix)
    df.to_excel(filename, index=False, header=False)


def highlight_diagonal(filename):
    wb = load_workbook(filename)
    ws = wb.active

    diagonal_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for i in range(1, min(ws.max_row, ws.max_column) + 1):
        cell = ws.cell(row=i, column=i)
        cell.fill = diagonal_fill

    wb.save(filename)
def create_heatmap(angle_matrix, sequence_names, filename):
    plt.figure(figsize=(10, 8))
    sns.heatmap(angle_matrix, xticklabels=sequence_names, yticklabels=sequence_names, cmap='coolwarm')
    plt.title('Angle Matrix Heatmap')
    plt.xlabel('Data')
    plt.ylabel('Data')
    # plt.xticks(rotation=90)
    # plt.yticks(rotation=0)

    plt.tight_layout()
    plt.savefig(filename)

    plt.show()
def create_interactive_heatmap(angle_matrix, sequence_names):
    fig = px.imshow(
        angle_matrix,
        labels=dict(x="Data", y="Data", color="Angle"),
        x=sequence_names,
        y=sequence_names,
        color_continuous_scale='RdBu'
    )
    
    fig.update_layout(
        title='256 Angle Matrix Heatmap',
        xaxis_nticks=len(sequence_names),
        yaxis_nticks=len(sequence_names)
    )
    
    # Save the interactive plot as an HTML file
    # pio.write_html(fig, file=filename, auto_open=True)
    
    # Show the interactive plot in the notebook or a web browser
    fig.show()



def plot_2d(pca_results, sequence_names, char_names):
    # Create a DataFrame for the Plotly scatter plot
    df = pd.DataFrame({
        'PCA1': pca_results[:, 0],
        'PCA2': pca_results[:, 1],
        'sequence': sequence_names,
        'type': ['char' if seq in char_names else 'sequence' for seq in sequence_names]
    })

    # Create the interactive plot
    fig = px.scatter(df, x='PCA1', y='PCA2', text='sequence', 
                     color='type',  # Use the 'type' column for color
                     labels={'PCA1': 'PCA Component 1', 'PCA2': 'PCA Component 2'},
                     title='PCA Plot of Sequence Embeddings (256)')

    # Update the traces
    fig.update_traces(marker=dict(size=12, opacity=0.8),
                      textposition='top center')

    # Update the layout for better presentation
    fig.update_layout(title_x=0.5, 
                      template='plotly_white', 
                      xaxis=dict(title='PCA Component 1', gridcolor='lightgrey'),
                      yaxis=dict(title='PCA Component 2', gridcolor='lightgrey'))

    fig.show()

def plot_cross_correlation(matrix):

    # Flatten the matrix
    flat_matrix = matrix.flatten()

    # Compute auto-correlation
    correlation = correlate(flat_matrix, flat_matrix, mode='full')

    # Compute lags
    lags = np.arange(-len(flat_matrix) + 1, len(flat_matrix))

    # Plot lag vs. correlation
    plt.figure(figsize=(10, 5))
    plt.plot(lags, correlation)
    plt.xlabel('Lag')
    plt.ylabel('Correlation')
    plt.title('Correlation vs. Lag')
    plt.grid(True)
    plt.show()

    return lags, correlation


# input_file_path = 'data/shakespeare_char/input_sorted.txt'
# output_file_path = 'data/shakespeare_char/input_sorted_with_chars.txt'  # Use a different output file to avoid overwriting
# input_list = []
single_characters = ['A', 'B', 'C', 'D']

# # Read the input file
# with open(input_file_path, 'r') as f:
#     data = f.readlines()
#     for line in data:
#         input_list.append(line.strip())

# # Append the single characters to the list
# input_list.extend(single_characters)

# # Sort the list
# sorted_list = sorted(input_list)

# # Write the sorted list to the output file
# with open(output_file_path, "w") as f:
#     for item in sorted_list:
#         f.write(item)
#         f.write('\n')


char_names = []
char_embeddings = []


for sequence in single_characters:
    sequence = sequence.strip()  # Remove leading/trailing whitespace
    char_names.append(sequence)
    start_ids = encode(sequence)

    x = torch.tensor(start_ids, dtype=torch.long, device=device)[None, ...]

    # Extract embeddings
    with torch.no_grad():
        token_embeddings = model.transformer.wte(x)  # Token embeddings
        position_ids = torch.arange(x.size(1), dtype=torch.long, device=x.device)
        position_embeddings = model.transformer.wpe(position_ids)  # Position embeddings
        embeddings = token_embeddings + position_embeddings  # Combine token and position embeddings

        # Flatten the embeddings and add to the list
        char_embeddings.append(embeddings.view(-1).cpu().numpy())
print(np.array(char_embeddings).shape)
# print(char_embeddings)


all_embeddings = []
sequence_names = []
input_file_path = 'data/shakespeare_char/full_dataset.txt'

with open(input_file_path, 'r') as f:
    data = f.readlines()
    for sequence in data:
        sequence = sequence.strip()  # Remove leading/trailing whitespace
        sequence_names.append(sequence)
        start_ids = encode(sequence)

        x = torch.tensor(start_ids, dtype=torch.long, device=device)[None, ...]

        # Extract embeddings
        with torch.no_grad():
            token_embeddings = model.transformer.wte(x)  # Token embeddings
            position_ids = torch.arange(x.size(1), dtype=torch.long, device=x.device)
            position_embeddings = model.transformer.wpe(position_ids)  # Position embeddings
            embeddings = token_embeddings + position_embeddings  # Combine token and position embeddings

            # Flatten the embeddings and add to the list
            all_embeddings.append(embeddings.view(-1).cpu().numpy())

# Convert list to numpy array``
# print(all_embeddings)
all_embeddings_np = np.array(all_embeddings)
# # print(all_embeddings_np[0])
# # print(all_embeddings_np[1])
# print(all_embeddings_np.shape)


pca = PCA(n_components=2, svd_solver = 'full')

pca_result = pca.fit_transform(all_embeddings_np)

pca = PCA(n_components=2, svd_solver = 'full')
char_result = pca.fit_transform(char_embeddings)
combined_result = np.vstack((pca_result, char_result))
print(combined_result.shape)


combined_names = sequence_names + char_names
angle_matrix = calculate_angle_matrix(pca_result)
plot_2d(combined_result, combined_names, char_names)

create_interactive_heatmap(angle_matrix, sequence_names)

lags, correlation = plot_cross_correlation(angle_matrix)
print("cross correlation:", correlation)

# # Example usage
# x_corr_value = cross_correlation_distance(pca_result)

# print("cross-correlation:", x_corr_value)

# with open('x_corr_values.txt', 'a') as f:
#     l = f"128,{x_corr_value}\n"
#     f.write(l)

# heatmap_filename = 'xcorr_matrix_heatmap.png'
# create_heatmap(xcorr_matrix, sequence_names, heatmap_filename)


# cross_corr_matrix = calculate_cross_correlation_matrix(all_embeddings_np)
# print("x-correlation", cross_corr_matrix)
# x_corr_flatten = np.Flatten
# heatmap_filename = 'xcorr_matrix_heatmap.png'
# create_heatmap(cross_corr_matrix, heatmap_filename)



# print(all_embeddings_np.shape)





# # Perform PCA
# pca = PCA(n_components=2)
# pca_result = pca.fit_transform(all_embeddings_np)

# # Create a DataFrame for the Plotly scatter plot
# df = pd.DataFrame({
#     'PCA1': pca_result[:, 0],
#     'PCA2': pca_result[:, 1],
#     'sequence': sequence_names
# })

# # Create the interactive plot
# fig = px.scatter(df, x='PCA1', y='PCA2', text='sequence', 
#                  labels={'PCA1': 'PCA Component 1', 'PCA2': 'PCA Component 2'},
#                  title='PCA Plot of Sequence Embeddings (2D)')

# # Update the traces
# fig.update_traces(marker=dict(size=12, opacity=0.8),
#                   textposition='top center')

# # Update the layout for better presentation
# fig.update_layout(title_x=0.5, 
#                   template='plotly_white', 
#                   xaxis=dict(title='PCA Component 1', gridcolor='lightgrey'),
#                   yaxis=dict(title='PCA Component 2', gridcolor='lightgrey'))

# # Show the interactive plot
# fig.show()

# # Perform PCA
# pca = PCA(n_components=3)
# pca_result = pca.fit_transform(all_embeddings_np)

# # Create a DataFrame for the Plotly scatter plot
# df = pd.DataFrame({
#     'PCA1': pca_result[:, 0],
#     'PCA2': pca_result[:, 1],
#     'PCA3': pca_result[:, 2],
#     'sequence': sequence_names
# })

# # Create the interactive 3D plot
# fig = px.scatter_3d(df, x='PCA1', y='PCA2', z='PCA3', text='sequence',
#                     labels={'PCA1': 'PCA Component 1', 'PCA2': 'PCA Component 2', 'PCA3': 'PCA Component 3'},
#                     title='3D PCA Plot of Sequence Embeddings')

# # Update the traces
# fig.update_traces(marker=dict(size=8, opacity=0.8),
#                   textposition='top center')

# # Update the layout for better presentation
# fig.update_layout(title_x=0.5, 
#                   template='plotly_white')

# # Show the interactive plot
# fig.show()


# cos_sim_matrix = cosine_similarity(all_embeddings_np)
# distance_matrix = np.round(1 - cos_sim_matrix, 2)


# # Create heatmap
# plt.figure(figsize=(10, 8))
# sns.heatmap(distance_matrix, xticklabels=sequence_names, yticklabels=sequence_names, cmap='coolwarm')
# plt.title('Cosine Similarity Heatmap of Synthetic Data Embeddings')
# plt.xlabel('Data')
# plt.ylabel('Data')
# # plt.xticks(rotation=90)
# # plt.yticks(rotation=0)
# plt.tight_layout()
# plt.show()
# # # run generation
# # with torch.no_grad():
# #     with ctx:
# #         for k in range(num_samples):
# #             y = model.generate(x, max_new_tokens, temperature=temperature, top_k=top_k)
# #             print(decode(y[0].tolist()))
# #             print('---------------')
